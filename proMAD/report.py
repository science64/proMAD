import csv
import io
import json
import os
import warnings
from datetime import datetime
from pathlib import Path
from tempfile import NamedTemporaryFile
import jinja2
from pylatexenc.latexencode import unicode_to_latex
import shutil

from typing import List, Dict

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as opImage
from openpyxl.styles import NamedStyle, Alignment, Font
from openpyxl.utils import get_column_letter

from proMAD import config


class Report:
    norm_descriptions = dict(
        raw="list of averages for all time-steps based on the original image",
        raw_bg="list of averages for all time-steps of the original image"
               " reduced by the histogram based background value",
        local_bg="mean of the ratios between the original images, and the extracted backgrounds",
        hist_fg="linear correlation between background (histogram) evolution to the average foreground value",
        hist_raw="linear correlation between background (histogram) evolution to the average original image",
        reac="estimate of the catalytic enzyme concentration"
    )

    norm_names = dict(
        raw='Spot average',
        raw_bg='Background corrected spot average',
        local_bg='Ratio to local background',
        hist_fg='Foreground to global background ratio',
        hist_raw='Ratio to global background',
        reac='Catalytic enzyme concentration'
    )

    norm_return = dict(
        raw=('values',),
        raw_bg=('values',),
        local_bg=('value', 'std'),
        hist_fg=('value', 'intercept', 'r_squared'),
        hist_raw=('value', 'intercept', 'r_squared'),
        reac=('value', 'std')
    )

    norm_unit = dict(
        raw='greyscale [0:1]',
        raw_bg='greyscale [0:1]',
        local_bg='background ratio',
        hist_fg='background ratio',
        hist_raw='background ratio',
        reac='concentration in mol/L'
    )

    tex_env = jinja2.Environment(
        block_start_string=r'\BLOCK{',
        block_end_string='}',
        variable_start_string=r'\VAR{',
        variable_end_string='}',
        comment_start_string=r'\#{',
        comment_end_string='}',
        line_statement_prefix='%%',
        line_comment_prefix='%#',
        trim_blocks=True,
        autoescape=False,
        lstrip_blocks=True,
        loader=jinja2.FileSystemLoader(str(config.template_folder.absolute()))
    )

    @classmethod
    def get_details(cls, aa, norm, tex=False):
        now = datetime.now()
        now = now.replace(microsecond=0)
        detail_values = dict(
            program='proMAD',
            version=config.version,
            url=config.url,
            date=now.date().isoformat(),
            time=now.time().isoformat(),
            array_name=aa.array_data['name'],
            array_type=aa.array_data['array_type'],
            array_id=aa.array_data['id'],
            norm_name=cls.norm_names[norm],
            norm=norm,
            norm_description=cls.norm_descriptions[norm],
            unit=cls.norm_unit[norm]
        )
        if tex:
            for key in detail_values:
                detail_values[key] = unicode_to_latex(detail_values[key]
                                                      )
        detail_names = dict(
            date='Date',
            time='Time',
            program='Program',
            version='Version',
            url='URL',
            array_name='Array Name',
            array_type='Array Type',
            array_id='Array ID',
            norm_name='Method',
            norm='Method key',
            norm_description='Method description',
            unit='Unit'
        )

        return detail_values, detail_names

    @classmethod
    def exp_excel(cls, aa, file, norm='hist_raw', additional_info=None):
        """
        Export results in an Excel file.

        Parameters
        ----------
        aa:
            ArrayAnalyse instant
        norm: str
                evaluation strategy selection (see ArrayAnalyse.evaluate)
        file:
            file can be a path to a file (a string), a path-like object, or a file-like object
        additional_info: List[Dict]
            list with dictionaries containing 'name' and 'value' key
        """

        if not aa.is_finalized:
            warnings.warn('Data collection needs to be finalized to generate a report.', RuntimeWarning)
            return None
        wb = Workbook()

        highlight = NamedStyle(name="highlight")
        highlight.font = Font(bold=True, size=20)
        wb.add_named_style(highlight)

        bold = NamedStyle(name="bold")
        bold.font = Font(bold=True)
        wb.add_named_style(bold)

        bold_right = NamedStyle(name="bold_right")
        bold_right.font = Font(bold=True)
        bold_right.alignment = Alignment(horizontal='right')
        wb.add_named_style(bold_right)

        bold_center = NamedStyle(name="bold_center")
        bold_center.font = Font(bold=True)
        bold_center.alignment = Alignment(horizontal='center')
        wb.add_named_style(bold_center)

        # Overview worksheet
        ws = wb.active

        alignment_img = opImage(config.template_folder/'logo.png')
        ws.add_image(alignment_img, f"B1")

        ws['E4'] = f'proMAD'
        ws['E5'] = config.version
        ws['E4'].style = highlight
        ws.title = "Overview"
        data = aa.evaluate(norm=norm)
        row_offset = 12
        column_offset = 2
        if cls.norm_return[norm] == ('values',):
            for entry in data:
                ws.cell(column=entry['position'][1] + column_offset,
                        row=entry['position'][0] + row_offset,
                        value=entry['value'][-1])
        else:
            for entry in data:
                ws.cell(column=entry['position'][1] + column_offset,
                        row=entry['position'][0] + row_offset,
                        value=entry['value'][0])

        for column in range(sum(aa.array_data['net_layout_x'])):
            ws.cell(column=column + column_offset, row=row_offset - 1, value=column + 1).style = bold_center

        for row in range(sum(aa.array_data['net_layout_y'])):
            ws.cell(column=column_offset - 1, row=row + row_offset, value=get_column_letter(row + 1)).style = bold_right

        alignment_img_raw = io.BytesIO()
        aa.figure_alignment(file=alignment_img_raw)
        alignment_img = opImage(alignment_img_raw)
        ws[f"B{sum(aa.array_data['net_layout_y']) + row_offset + 1}"] = 'Fig. 1: Overview and alignment check'
        ws.add_image(alignment_img, f"B{sum(aa.array_data['net_layout_y'])+row_offset+2}")

        # Result worksheet
        ws = wb.create_sheet()
        ws.title = "Results"
        new_rows = []
        ref_spots = []
        data = aa.evaluate(norm=norm, double_spot=True)
        for entry in data:
            pos = aa.get_position_string(entry['position'])
            if not isinstance(pos, str):
                pos = ", ".join(pos)
            if 'Reference' in entry['info'][0] or 'Negative Controls' in entry['info'][0]:
                ref_spots.append([entry['info'][0], pos, entry['value']])
            else:
                new_rows.append([entry['info'][0], pos, entry['value']])

        ws['A1'] = 'Membrane Results'
        ws['A1'].style = highlight
        ws.append([])
        ws.append(('Name', 'Position', 'Value'))
        ws['A3'].style = bold
        ws['B3'].style = bold
        ws['C3'].style = bold_right

        data_start = 5 + len(ref_spots)
        for row in sorted(ref_spots, key=lambda s: s[2], reverse=True):
            ws.append(row)
        ws.append([])
        for row in sorted(new_rows, key=lambda s: s[2], reverse=True):
            ws.append(row)
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['C'].width = 12

        cutoff = max(min(len(new_rows), 15), int(len(new_rows)*0.2))
        values = Reference(ws, min_col=3, min_row=data_start, max_col=3, max_row=cutoff+data_start)
        categories = Reference(ws, min_col=1, min_row=data_start, max_col=1, max_row=cutoff+data_start)
        chart = BarChart()
        chart.add_data(values)
        chart.set_categories(categories)
        chart.y_axis.title = cls.norm_unit[norm][0].upper() + cls.norm_unit[norm][1:]
        chart.title = f'Top spots'
        chart.width = 30
        chart.height = 15
        chart.gapWidth = 25
        chart.legend = None
        ws.add_chart(chart, "E3")

        # Result details worksheet
        ws = wb.create_sheet()
        ws.title = "Result details"
        new_rows = []

        data = aa.evaluate(norm=norm)
        for entry in data:
            new_rows.append([entry['info'][0], aa.get_position_string(entry['position'])] + [v for v in entry['value']])
        ws['A1'] = 'Membrane Result Details'
        ws['A1'].style = highlight
        ws.append([])
        ws.append(['Name', 'Position'] + [head.title() for head in cls.norm_return[norm]])
        ws['A3'].style = bold
        ws['B3'].style = bold
        for n in range(len(cls.norm_return[norm])):
            ws[get_column_letter(n+3)+'3'].style = bold_right
        for row in sorted(new_rows, key=lambda s: s[0].lower()):
            ws.append(row)

        ws.column_dimensions['A'].width = 20
        for n in range(len(new_rows[0])-2):
            ws.column_dimensions[get_column_letter(n+3)].width = 12

        # Info worksheet
        ws = wb.create_sheet()
        ws.title = "Info"

        ws['A1'] = f'Technical data'
        ws['A1'].style = highlight
        detail_values, detail_names = cls.get_details(aa, norm)
        info_list = ['date', 'time', 'skip',
                     'program', 'version', 'url', 'skip',
                     'array_name', 'array_type', 'array_id', 'norm_name', 'norm', 'norm_description', 'unit'
                     ]
        tech_data_list = []
        for key in info_list:
            if key == 'skip':
                tech_data_list.append(('', ''))
            else:
                tech_data_list.append((detail_names[key], detail_values[key]))

        if additional_info:
            tech_data_list += [('', '')] + [(entry['name'], entry['value']) for entry in additional_info]

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        for n, content in enumerate(tech_data_list):
            ws.cell(column=1, row=n+3, value=content[0]).style = bold_right
            ws.cell(column=2, row=n+3, value=content[1])

        if isinstance(file, os.PathLike) or isinstance(file, str):
            wb.save(file)
        elif isinstance(file, (io.RawIOBase, io.BufferedIOBase)):
            with NamedTemporaryFile() as tmp:
                wb.save(tmp.name)
                tmp.seek(0)
                file.write(tmp.read())

    @classmethod
    def exp_csv(cls, aa, file, norm='hist_raw', additional_info=None):
        """
        Export results in a comma separated file.

        Parameters
        ----------
        aa:
            ArrayAnalyse instant
        norm: str
            evaluation strategy selection (see ArrayAnalyse.evaluate)
        file:
            file can be a path to a file (a string), a path-like object, or a file-like object (string based)
        additional_info:
            is ignored for CSV
        """

        data = aa.evaluate(norm=norm)
        header = ['Name', 'Position'] + [head.title() for head in cls.norm_return[norm]]

        if isinstance(file, os.PathLike) or isinstance(file, str):
            with open(file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerows((header, ))
                writer.writerows([[row['info'][0], row['position']] + [f'{v:.8g}' for v in row['value']] for row in data])

        elif isinstance(file, (io.TextIOBase, io.TextIOWrapper)):
            writer = csv.writer(file)
            writer.writerows((header,))
            writer.writerows([[row['info'][0], row['position']] + [f'{v:.8g}' for v in row['value']] for row in data])

    @classmethod
    def exp_latex(cls, aa, file, norm='hist_raw', additional_info=None):
        """
        Export results as tex file.

        Parameters
        ----------
        aa:
            ArrayAnalyse instant
        norm: str
                evaluation strategy selection (see ArrayAnalyse.evaluate)
        file:
            can be a path to a file (a string), a path-like object, or a file-like object (string based)
        additional_info: List[Dict]
            list with dictionaries containing name and value key
        """

        overview = []
        reference = []
        data = aa.evaluate(norm=norm, double_spot=True)
        for entry in data:
            pos = aa.get_position_string(entry['position'])
            if not isinstance(pos, str):
                pos = ", ".join(pos)

            if 'Reference' in entry['info'][0]:
                reference.append(dict(
                    name=unicode_to_latex(entry['info'][0]),
                    gene_id=entry['info'][1],
                    position=pos,
                    value=f"{entry['value']:.4g}",
                    sort=entry['value'])
                )
            else:
                overview.append(dict(
                    name=unicode_to_latex(entry['info'][0]),
                    gene_id=entry['info'][1],
                    position=pos,
                    value=f"{entry['value']:.4g}",
                    sort=entry['value'])
                )

        best = sorted(overview, key=lambda s: s['sort'], reverse=True)[:15]
        overview = sorted(overview+reference, key=lambda s: s['name'].lower())
        info_dict = dict(overview=overview, best=best)
        info_dict['dv'], info_dict['dn'] = cls.get_details(aa, norm, tex=True)

        col_num = sum(aa.array_data['net_layout_x'])
        row_num = sum(aa.array_data['net_layout_y'])

        if additional_info:
            for entry in additional_info:
                entry['name'] = unicode_to_latex(entry['name'])
                entry['value'] = unicode_to_latex(entry['value'])
        info_dict['additional_info'] = additional_info
        info_dict['ai'] = dict(
            row=[(get_column_letter(row_num-n), f'{n/row_num + 0.5/row_num:.3f}') for n in range(row_num)],
            col=[(str(n+1), f'{n/col_num + 0.5/col_num:.3f}') for n in range(col_num)]
        )

        template = cls.tex_env.get_template('short_report.tex')

        content = template.render(**info_dict)

        if isinstance(file, os.PathLike) or isinstance(file, str):
            file = Path(file)
            file.write_text(content)
            shutil.copy(config.template_folder / 'logo.png', file.parent / 'logo.png')
            aa.figure_alignment(file=file.parent / 'figure_alignment.jpg')

        elif isinstance(file, (io.TextIOBase, io.TextIOWrapper)):
            file.seek(0)
            file.write(content)

    @classmethod
    def exp_json(cls, aa, file, norm='hist_raw', additional_info=None):
        """
        Export results in a structured json file.

        Parameters
        ----------
        aa:
            ArrayAnalyse instant
        norm: str
                evaluation strategy selection (see ArrayAnalyse.evaluate)
        file:
            can be a path to a file (a string), a path-like object, or a file-like object (string based)
        additional_info: List[Dict]
            list with dictionaries containing 'key' and 'value' key
        """

        data = aa.evaluate(norm=norm)

        if len(cls.norm_return[norm]) == 1:
            for entry in data:
                value = entry['value']
                del entry['value']
                entry[cls.norm_return[norm][0]] = [float(v) for v in value]
        else:
            for entry in data:
                value = entry['value']
                del entry['value']
                for n, name in enumerate(cls.norm_return[norm]):
                    entry[name] = float(value[n])

        export_dict = dict(
            result=data,
            info=cls.get_details(aa, norm)[0]
        )

        if additional_info:
            info_update = {entry['key']: entry['value'] for entry in additional_info}
        else:
            info_update = {}

        if additional_info:
            export_dict['info'].update(info_update)

        if isinstance(file, os.PathLike) or isinstance(file, str):
            Path(file).write_text(json.dumps(export_dict, indent=4))

        elif isinstance(file, (io.TextIOBase, io.TextIOWrapper)):
            file.seek(0)
            file.write(json.dumps(export_dict, indent=4))
