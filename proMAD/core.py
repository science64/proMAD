import json
import os
import re
import io
import string
import warnings
from collections import Counter
from pathlib import Path
import tarfile

import matplotlib.pyplot as plt
import numpy as np
from scipy import stats
from scipy.optimize import curve_fit, minimize_scalar
from skimage import io as ski_io
from skimage import img_as_float, img_as_ubyte
from skimage import measure
from skimage import transform
from skimage.color import rgb2gray
from skimage.external import tifffile
from skimage.morphology import reconstruction
from skimage.transform import rotate, rescale
from skimage.util import invert
from PIL import Image
from openpyxl.utils import get_column_letter

from proMAD import Report, config


class ArrayAnalyse(object):
    """
    Analyse a set of micro arrays with variable exposure.

    Notes
    -----

    To enable an more detailed output *debug* can be set to 'basic' or 'plot'.

    >>> aa = ArrayAnalyse('ARY022B')
    >>> aa.debug = 'plot'
    >>> aa.debug = 'basic'

    This program will test images to ensure they match certain expected properties.
    If one of the tests are failed the image is not added to the collection.
    To disable this function set *strict* to False.

    >>> aa.strict = False

    To modify the test the conditions can be modified.
    The following parameters are available:

    - `aa.test_warp_length_ration`: deviation from the x/y ratio (0.1)
    - `aa.test_warp_angle`: deviation from the angle in radians (0.05)
    - `aa.test_warp_distance`: allowed (distance/short edge) between guessed anchor
       positions and nearest found contour (0.275)
    - `aa.test_quality_resolution`: minimal used portion of the brightness spectra (0.1)
    - `aa.test_quality_exposure`: maximal allowed number of full exposed pixels (1)

    """
    
    def __init__(self, array_type, silent=False):
        """
        Initialize ArrayAnalyse object

        Notes
        -----
        You can list all available options for `array_type` with *list_types()*

        >>> ArrayAnalyse.list_types()
        ARY022B
            Name: Human XL Cytokine Array Kit
            Type: Proteome Profiler Array
            Company: R&D Systems, Inc
            Source: https://resources.rndsystems.com/pdfs/datasheets/ary022b.pdf
        ...


        Parameters
        ----------
        array_type: str
            id of the used micro array type
        silent: bool
            be silent if True
        """

        if not silent:
            def verbose_print(*args):
                print(*args)
        else:
            def verbose_print(*args):
                pass

        self.silent = silent
        self.verbose_print = verbose_print
        self.array_types = self.get_types()
        self.array_data = self.array_types[array_type]
        self.source_images = []
        self.raw_images = []
        self.backgrounds = []
        self.foregrounds = []
        self.bg_parameters = []
        self.original_average = []
        self.original_names = []
        self.original_index = []
        self.raw_index = []
        self.exposure = []
        self.meta_data = []
        self.debug = None
        self.strict = True

        # Reaction fit parameters
        self.start_time = 60  # s
        self.k_reac = 1.4E6  # 1/(mol/L * s)
        self.kappa_fit_count = 10
        self._kappa = 1
        self._fit_selection = []

        # Test parameters
        self.test_warp_length_ration = 0.1
        self.test_warp_angle = 0.05
        self.test_warp_distance = 0.275
        self.test_quality_resolution = 0.1
        self.test_quality_exposure = 1

        self.is_finalized = False
        self.has_exposure = False

        self.save_list_base = ["meta_data",
                               "start_time",
                               "k_reac",
                               "kappa_fit_count",
                               "test_warp_length_ration",
                               "test_warp_angle",
                               "test_warp_distance",
                               "test_quality_resolution",
                               "test_quality_exposure",
                               "debug",
                               "strict",
                               "has_exposure",
                               "is_finalized",
                               "_kappa"]
        self.save_list_data = ['source_images', 'raw_images', 'backgrounds', 'foregrounds',
                               'bg_parameters', 'original_index', 'original_average',
                               'original_average', 'raw_index']

        self.grid_position = np.zeros((sum(self.array_data['net_layout_x']), sum(self.array_data['net_layout_y']), 2))

        px, py = 0, 0
        for xc, x_block in enumerate(self.array_data['net_layout_x']):
            for x in range(x_block):
                self.grid_position[px, :, 0] = (config.scale / 2.0 +
                                                (px + xc * self.array_data['net_skip_size']) * config.scale)
                px += 1

        for yc, y_block in enumerate(self.array_data['net_layout_y']):
            for y in range(y_block):
                self.grid_position[:, py, 1] = (config.scale / 2.0 +
                                                (py + yc * self.array_data['net_skip_size']) * config.scale)
                py += 1

    @staticmethod
    def _angle_between(v1, v2):
        v1_u = v1 / np.linalg.norm(v1)
        v2_u = v2 / np.linalg.norm(v2)
        return np.arccos(np.clip(np.dot(v1_u, v2_u), -1.0, 1.0))

    @staticmethod
    def _compare_version(a, b):
        """
        Return True if version b is the same or higher than version a

        Parameters
        ----------
        a: array_like
        b: array_like
        """
        return a[0] < b[0] or (a[0] <= b[0] and a[1] < b[1]) or (a[0] <= b[0] and a[1] <= b[1] and a[2] <= b[2])

    @staticmethod
    def _failed_passed(value):
        if value:
            return '✅'
        else:
            return '❌'

    @staticmethod
    def _add_2d_tuple(a, b):
        return a[0] + b[0], a[1] + b[1]

    @staticmethod
    def get_types():
        """
        Return all available array types.

        Returns
        -------
        dict
            Dictionary with the all array data

        """

        array_types = dict()
        for file_path in config.array_data_folder.iterdir():
            if file_path.is_file():
                if file_path.suffix == '.json':
                    new_array = json.loads(file_path.read_text())
                    array_types[new_array['id']] = new_array
        return array_types

    @classmethod
    def list_types(cls):
        """
        List all available array types.

        """
        array_types = cls.get_types()
        for array_id, array in array_types.items():
            print(array_id)
            print(f'\tName: {array["name"]}')
            print(f'\tType: {array["array_type"]}')
            print(f'\tCompany: {array["company"]}')
            print(f'\tSource: {array["source"]}\n')

    @property
    def raw_names(self):
        return [self.original_names[i] for i in self.raw_index]

    def save(self, file):
        """
        Saves the finalized content of an ArrayAnalyse instant into a .tar file

        Parameters
        ----------
        file:
            can be a path to a file (a string), a path-like object, or a file-like object

        """
        if not self.is_finalized:
            warnings.warn('Data collection needs to be finalized to save.', RuntimeWarning)
            return None
        base_data = dict(array_type=self.array_data['id'],
                         silent=self.silent,
                         version=config.version_number)
        for entry in self.save_list_base:
            base_data[entry] = getattr(self, entry)

        data_file = io.BytesIO()
        if self.has_exposure:
            np.savez_compressed(data_file,
                                source_images=self.source_images,
                                raw_images=self.raw_images,
                                backgrounds=self.backgrounds,
                                foregrounds=self.foregrounds,
                                bg_parameters=self.bg_parameters,
                                exposure=self.exposure,
                                original_names=self.original_names,
                                original_index=self.original_index,
                                original_average=self.original_average,
                                raw_index=self.raw_index,
                                _fit_selection=self._fit_selection
                                )
        else:
            np.savez_compressed(data_file,
                                source_images=self.source_images,
                                raw_images=self.raw_images,
                                backgrounds=self.backgrounds,
                                foregrounds=self.foregrounds,
                                bg_parameters=self.bg_parameters,
                                original_names=self.original_names,
                                original_index=self.original_index,
                                original_average=self.original_average,
                                raw_index=self.raw_index,
                                )

        if isinstance(file, os.PathLike) or isinstance(file, str):
            tar = tarfile.open(file, mode="w")
        elif isinstance(file, (io.RawIOBase, io.BufferedIOBase)):
            tar = tarfile.open(fileobj=file, mode="w")
        else:
            raise TypeError(f'Type {type(file)} not supported to save the data.')

        base = json.dumps(base_data).encode("utf-8")
        ti = tarfile.TarInfo("base.json")
        ti.size = len(base)
        tar.addfile(ti, io.BytesIO(base))

        ti = tarfile.TarInfo("data.npz")
        ti.size = data_file.tell()
        data_file.seek(0)
        tar.addfile(ti, data_file)

        tar.close()

    @classmethod
    def load(cls, file):
        """
        Load the content of a saved ArrayAnalyse instant from file

        Parameters
        ----------
        file:
            can be a path to a file (a string), a path-like object, or a file-like object

        Returns
        -------
        ArrayAnalyse
            returns a ArrayAnalyse instance
        """
        if isinstance(file, os.PathLike) or isinstance(file, str):
            tar = tarfile.TarFile.open(file)
        elif isinstance(file, (io.RawIOBase, io.BufferedIOBase)):
            tar = tarfile.TarFile.open(fileobj=file)
        else:
            raise TypeError(f'Type {type(file)} not supported to load.')
        base_data = None
        data = None
        for member in tar.getmembers():
            if member.name == 'base.json':
                base_data = json.loads(tar.extractfile(member).read().decode('utf-8'))
            if member.name == "data.npz":
                data = np.load(tar.extractfile(member))

        if base_data is None or data is None:
            tar.close()
            raise TypeError("The loaded save file was not valid.")

        if not cls._compare_version(config.allowed_load_version, base_data['version']):
            version_str = "{}.{}.{}".format(*base_data["version"])
            tar.close()
            raise TypeError(f'A save file from version {version_str} cannot be loaded.')
        aa = cls(base_data['array_type'], silent=base_data['silent'])
        for name in aa.save_list_base:
            setattr(aa, name, base_data[name])
        if aa.has_exposure:
            aa.save_list_data += ['exposure', '_fit_selection']
        for name in aa.save_list_data:
            setattr(aa, name, data[name].copy())
        tar.close()

        return aa

    def reset_collection(self):
        """
        Empties the image collections

        """
        self.source_images = []
        self.raw_images = []
        self.backgrounds = []
        self.foregrounds = []
        self.bg_parameters = []
        self.meta_data = []
        self.exposure = []
        self.original_average = []
        self.original_index = []
        self.original_names = []
        self.raw_index = []
        self.original_names = []
        self.is_finalized = False
        self.has_exposure = False

    def load_collection(self, data_input, rotation=None, finalize=True):
        """
        Load multiple files into the collection.
        Includes just files that are directly in the folders listed in *data_input* and does not search subfolders.

        Parameters
        ----------
        data_input: list(str) or list(Path) or Path or str
            a single input folder or list of folders / files
        rotation: int or float
            apply a rotation to all images
        finalize: bool
            defines if a collection should be directly finalized

        """

        if self.is_finalized:
            warnings.warn("Data is already finalized. You can use reset_collection to start over.", RuntimeWarning)
            return None

        if isinstance(data_input, str):
            input_paths = [Path(data_input)]
        elif isinstance(data_input, Path):
            input_paths = [data_input]
        else:
            input_paths = []
            for input_entry in data_input:
                input_paths.append(Path(input_entry))
        input_file_paths = []
        for input_path in input_paths:
            if input_path.is_dir():
                for folder_file_path in input_path.iterdir():
                    if folder_file_path.is_file():
                        input_file_paths.append(folder_file_path)
            elif input_path.is_file():
                input_file_paths.append(input_path)
        if not input_file_paths:
            warnings.warn('Found no images', RuntimeWarning)
            return None
        for file_path in input_file_paths:
            if file_path.name[0] == '.':
                continue
            self.load_image(file_path, rotation=rotation)

        if finalize and self.raw_images:
            self.finalize_collection()
        else:
            self.reset_collection()

    def finalize_collection(self):
        """
        Orders the image collections and convert them into numpy arrays.
        The ordering is either derived from the background intensity extracted
        from a histogram or exposure time in the image metadata.

        """
        if self.is_finalized:
            warnings.warn("Data is already finalized. You can use reset_collection to start over.", RuntimeWarning)
            return None

        if len(self.raw_images) == 0:
            warnings.warn("There is no data to be finalized.", RuntimeWarning)
            return None

        self.bg_parameters = np.array(self.bg_parameters)
        self.raw_index = np.array(self.raw_index)
        self.original_average = np.array(self.original_average)
        self.exposure = np.array(self.exposure)

        if self.exposure.size == self.bg_parameters.size:
            order = np.argsort(self.exposure)
            self.exposure = self.exposure[order]
            self.has_exposure = True
        else:
            order = np.argsort(self.bg_parameters)
            self.exposure = []

        orginal_order = np.argsort(self.original_average)
        raw_images_array = np.zeros(shape=(self.raw_images[0].shape + (len(order),)),
                                    dtype=self.raw_images[0].dtype)
        backgrounds_array = np.zeros(shape=(self.backgrounds[0].shape + (len(order),)),
                                     dtype=self.backgrounds[0].dtype)
        foregrounds_array = np.zeros(shape=(self.foregrounds[0].shape + (len(order),)),
                                     dtype=self.foregrounds[0].dtype)
        source_images = []
        meta = []
        for n, i in enumerate(order):
            raw_images_array[:, :, n] = self.raw_images[i]
            backgrounds_array[:, :, n] = self.backgrounds[i]
            foregrounds_array[:, :, n] = self.foregrounds[i]
            source_images.append(self.source_images[i])
            meta.append(self.meta_data[i])
        self.raw_images = raw_images_array
        self.backgrounds = backgrounds_array
        self.foregrounds = foregrounds_array
        self.bg_parameters = self.bg_parameters[order]
        self.raw_index = self.raw_index[order]
        self.original_index = orginal_order
        self.original_average = self.original_average[orginal_order]
        self.meta_data = meta
        self.source_images = source_images
        self.is_finalized = True
        if self.has_exposure:
            self.minimize_kappa()

        if self.debug == 'plot':  # pragma: no cover
            self.figure_alignment()
            self.figure_contact_sheet()

    def modify_exposure(self, exposure_info, test=False):
        """
        Add or modify exposure information of a finalized collection.

        Notes
        -----
        In case set exposure time changes the order

        Parameters
        ----------
        exposure_info: Union[dict, list]
            either list of exposure times with the same length and order as shown by *raw_names* or
            dict describing start and step size of the exposure `{'start': 10, 'step': 30}`
            unit in seconds
        test: bool
            if True no changes are made

        Returns
        -------
        exposure: list
            generated list of exposure in order of *raw_names*
        reorder: bool
            if the new exposure implies a reordering True is returned

        """

        if not self.is_finalized:
            warnings.warn('Data collection needs to be finalized to amend exposure information.', RuntimeWarning)
            return None
        if isinstance(exposure_info, dict):
            if 'start' in exposure_info and 'step' in exposure_info:
                raw_exposure = exposure_info['start'] + np.array(range(len(self.original_average))) * exposure_info['step']
                raw_exposure_lookup = {i: raw_exposure[n] for n, i in enumerate(self.original_index)}
                exposure = [raw_exposure_lookup[i] for i in self.raw_index]
            else:
                warnings.warn('Exposure information needs the keys "start" and "step".', RuntimeWarning)
                return None, None
        else:
            if len(exposure_info) == len(self.bg_parameters):
                exposure = exposure_info
            else:
                warnings.warn('Exposure information has the wrong length.', RuntimeWarning)
                return None, None

        order = np.argsort(exposure)
        reorder = not np.all(order[:-1] <= order[1:])
        if test:
            return exposure, reorder

        self.exposure = exposure
        if reorder:
            raw_images_array = np.zeros_like(self.raw_images)
            backgrounds_array = np.zeros_like(self.backgrounds)
            foregrounds_array = np.zeros_like(self.foregrounds)
            source_images = []
            meta = []
            for n, i in enumerate(order):
                raw_images_array[:, :, n] = self.raw_images[:, :, i]
                backgrounds_array[:, :, n] = self.backgrounds[:, :, i]
                foregrounds_array[:, :, n] = self.foregrounds[:, :, i]
                source_images.append(self.source_images[i])
                meta.append(self.meta_data[i])
            self.raw_images = raw_images_array
            self.backgrounds = backgrounds_array
            self.foregrounds = foregrounds_array
            self.bg_parameters = self.bg_parameters[order]
            self.raw_index = self.raw_index[order]
            self.meta_data = meta
            self.source_images = source_images

        self.has_exposure = True
        self.minimize_kappa()
        return self.exposure, reorder

    def load_image(self, file, rotation=None, filename=None, meta_data=None):
        """
        Load a single image file into the collection.


        Parameters
        ----------
        file:
            file can be a path to a file (a string or path-like object), a file-like object, or a numpy ndarray;
            if file is None result is directly shown
        rotation: int or float
            apply a rotation to the images
        filename: str
            if a file-like object is submitted the filename is needed for type identification (".tif", ".png", ...)
        meta_data: dict()


        """

        if self.is_finalized:
            warnings.warn("Data is already finalized. You can use reset_collection to start over.", RuntimeWarning)
            return None

        file_system = True
        if isinstance(file, os.PathLike) or isinstance(file, str):
            self.verbose_print(f'Load image: {file}')
            file = Path(file)
            source_image = ski_io.imread(file.absolute(), plugin='imageio')
            filename = file.name
        elif isinstance(file, np.ndarray):
            source_image = file
        elif isinstance(file, (io.RawIOBase, io.BufferedIOBase)):
            source_image = ski_io.imread(file, plugin='imageio')
            file_system = False
        else:
            warnings.warn(f'Type {type(file)} not supported for read image. (skipped)', RuntimeWarning)
            return None

        if meta_data is None:
            if filename.lower().endswith('.tif'):
                if file_system:
                    with tifffile.TiffFile(str(file.absolute())) as tif_data:
                        tags = [page.tags for page in tif_data.pages]
                else:
                    file.seek(0)  # imread does not rewind file properly
                    with tifffile.TiffFile(file) as tif_data:
                        tags = [page.tags for page in tif_data.pages]
                if tags:
                    try:
                        meta_data = json.loads(tags[0]['image_description'].value)
                    except (KeyError, json.decoder.JSONDecodeError):
                        pass

        if source_image.ndim == 3:
            warnings.warn("Load RGB image. Converting to greyscale alters the result. Use raw greyscale if possible.",
                          RuntimeWarning)
            source_image = rgb2gray(source_image)
            # if min value is black invert image
            if np.average(source_image) > np.max(source_image)//2:
                source_image = invert(source_image)
        if rotation:
            source_image = rotate(source_image, rotation, resize=True)

        source_image = img_as_float(source_image)
        self.original_average.append(np.average(source_image))
        if filename:
            self.original_names.append(filename)
        else:
            warnings.warn("No filename was given. This limits some functions.", RuntimeWarning)

        raw_image = self.warp_image(source_image, rotation=rotation)

        if raw_image is not None:
            used_resolution = np.max(raw_image) - np.min(raw_image)
            over_exposed = np.sum(raw_image >= 1.0)
            if self.debug:
                self.verbose_print('Quality checks')
                self.verbose_print(f'\tused resolution: {self._failed_passed(not (np.isnan(used_resolution)) and not (used_resolution < self.test_quality_resolution))} ({used_resolution*100:.1f}%)')
                self.verbose_print(f'\tfull exposed pixels: {self._failed_passed(not (np.isnan(over_exposed)) and not (over_exposed > self.test_quality_exposure))} ({over_exposed})')

            if self.strict:
                if (np.isnan(used_resolution) or np.isnan(over_exposed) or
                        (used_resolution < self.test_quality_resolution) or
                        (over_exposed > self.test_quality_exposure)):
                    self.verbose_print('\tImage skipped - quality check failed.')
                    return None

            self.source_images.append(source_image)
            self.raw_images.append(raw_image)
            seed = np.copy(raw_image)
            seed[1:-1, 1:-1] = raw_image.max()
            mask = raw_image
            filled = reconstruction(seed, mask, method='erosion')
            seed = np.copy(filled)
            seed[1:-1, 1:-1] = filled.min()
            mask = filled
            dilated = reconstruction(seed, mask, method='dilation')
            image = filled - dilated
            self.backgrounds.append(dilated)
            self.foregrounds.append(image)
            self.bg_parameters.append(self.background_histogram(raw_image))
            self.meta_data.append(meta_data)
            self.raw_index.append(len(self.original_average) - 1)
            if meta_data:
                if 'exposure_time' in meta_data:
                    self.exposure.append(meta_data['exposure_time'])

    def warp_image(self, image, rotation=None):
        """
        Warp images into a defined shape based on a contour search.


        Parameters
        ----------
        image: array_like
            docs_source image
        rotation: int or float
            apply a rotation to the images

        Returns
        -------
        array_like
            warped image (64bit float [0-1])

        """
        x_length = (sum(self.array_data['net_layout_x']) - 1 + (len(self.array_data['net_layout_x']) - 1) *
                    self.array_data['net_skip_size']) * config.scale
        y_length = (sum(self.array_data['net_layout_y']) - 1 + (len(self.array_data['net_layout_y']) - 1) *
                    self.array_data['net_skip_size']) * config.scale

        if rotation is None:
            rotation = 0

        # find bright spots and calculate their center
        contours = measure.find_contours(image, np.max(image) * 0.60)
        x = []
        y = []
        for n, contour in enumerate(contours):
            y.append(sum(contour[:, 0]) / len(contour))
            x.append(sum(contour[:, 1]) / len(contour))

        x = np.array(x)
        y = np.array(y)

        # TODO: use array_data to generate the initial guesses

        # try/except if no contours could be found
        try:
            ref_1_guess = (min(x), min(y))
            ref_2_guess = (min(x), max(y))
            ref_3_guess = (max(x), min(y))
        except ValueError:
            self.verbose_print(
                '\tImage skipped - position guess failed - not enough contours found.')
            return None

        ref_1_idx = np.argsort((x - ref_1_guess[0]) ** 2 + (y - ref_1_guess[1]) ** 2)[0]
        ref_2_idx = np.argsort((x - ref_2_guess[0]) ** 2 + (y - ref_2_guess[1]) ** 2)[0]
        ref_3_idx = np.argsort((x - ref_3_guess[0]) ** 2 + (y - ref_3_guess[1]) ** 2)[0]

        # test if found points make sense
        y_length_test = np.sqrt((x[ref_2_idx] - x[ref_1_idx]) ** 2 + (y[ref_2_idx] - y[ref_1_idx]) ** 2)
        x_length_test = np.sqrt((x[ref_2_idx] - x[ref_3_idx]) ** 2 + (y[ref_2_idx] - y[ref_3_idx]) ** 2)
        length_test = (x_length_test/y_length_test) / (x_length/y_length)

        angle_test = self._angle_between(np.array([x[ref_3_idx] - x[ref_1_idx], y[ref_3_idx] - y[ref_1_idx]]),
                                         np.array([x[ref_2_idx] - x[ref_1_idx], y[ref_2_idx] - y[ref_1_idx]]))/np.pi*2

        dis_1 = np.min(np.sqrt((x - ref_1_guess[0]) ** 2 + (y - ref_1_guess[1]) ** 2))
        dis_2 = np.min(np.sqrt((x - ref_2_guess[0]) ** 2 + (y - ref_2_guess[1]) ** 2))
        dis_3 = np.min(np.sqrt((x - ref_3_guess[0]) ** 2 + (y - ref_3_guess[1]) ** 2))
        dis_test = max(dis_1, dis_2, dis_3)/min(y_length_test, x_length_test)

        if self.debug == 'plot':  # pragma: no cover
            fig, ax = plt.subplots()
            ax.imshow(image, interpolation='nearest', cmap=plt.cm.gray)
            ax.scatter(x, y, zorder=2, marker='o', c=np.arange(len(x))+1, cmap=plt.cm.viridis, label='detected contours')
            ax.scatter((x[ref_1_idx], x[ref_2_idx], x[ref_3_idx]),
                       (y[ref_1_idx], y[ref_2_idx], y[ref_3_idx]),
                       zorder=4, color='red', marker='+', label='selected positions')
            ax.scatter((min(x), min(x), max(x)), (min(y), max(y), min(y)),
                       zorder=5, color='yellow', marker='x', label='guessed positions')
            ax.legend(bbox_to_anchor=(0., 1.02, 1., .102), loc='lower left',
                      ncol=3, mode="expand", borderaxespad=0.)
            fig.show()

        if self.debug:
            self.verbose_print('Warp checks')
            self.verbose_print(f'\tlength ratio: {self._failed_passed(not (np.isnan(length_test)) and not (abs(length_test-1.0) > self.test_warp_length_ration))} ({length_test:.4f})')
            self.verbose_print(f'\tangle: {self._failed_passed(not (np.isnan(angle_test)) and not (abs(angle_test-1.0) > self.test_warp_angle))} ({angle_test*90:.2f})')
            self.verbose_print(f'\tguess distance: {self._failed_passed(not (np.isnan(dis_test)) and not (dis_test > self.test_warp_distance))} ({dis_test:.4f})')

        if self.strict:
            if (np.isnan(length_test) or np.isnan(angle_test) or np.isnan(dis_test)
                    or abs(length_test-1.0) > self.test_warp_length_ration
                    or abs(angle_test-1.0) > self.test_warp_angle or dis_test > self.test_warp_distance):
                self.verbose_print('\tImage skipped - position guess failed - image maybe not rotated correctly at %i degrees,\n'
                                   '\t\tor the contrast is not good enough for the feature extraction.' % rotation)
                return None

        # define the reference spots locations and transform the image
        src = np.float32([[x[ref_1_idx], y[ref_1_idx]],
                          [x[ref_2_idx], y[ref_2_idx]],
                          [x[ref_3_idx], y[ref_3_idx]]])
        dst = np.float32([[config.scale, config.scale],
                          [config.scale, y_length + config.scale],
                          [config.scale + x_length, config.scale]])
        tform = transform.estimate_transform('affine', dst, src)
        warped = transform.warp(image, tform, output_shape=[int(y_length + config.scale * 2),
                                                            int(x_length + config.scale * 2)])

        return warped

    @staticmethod
    def background_histogram(raw_image, full=False):
        """
        Generate the histogram and find background split parameter

        Parameters
        ----------
        raw_image: array_like
            input image

        full: bool
            if True return also derivation

        """
        image_reduced = np.round(np.array(list(raw_image.flat))*2**11).astype('uint16')
        histo_counter = Counter(image_reduced)
        peak = histo_counter.most_common(1)[0][0]
        background_data = image_reduced[image_reduced < (int(peak * 2))]
        background_parameter = stats.norm.fit(background_data, loc=peak)
        if not full:
            return background_parameter[0]/(2**11)
        else:
            return background_parameter

    @staticmethod
    def get_position_string(position):
        """
        Convert coordinates to position string.

        Parameters
        ----------
        position: (int, int) or str or list
            coordinates  [(3,5), (8,12), ...]

        Returns
        -------
        str
            position string

        """
        if isinstance(position, str):
            return position
        if len(position) == 2:
            if isinstance(position[0], int) and isinstance(position[1], int):
                return get_column_letter(position[0]+1)+str(position[1]+1)

        result = []
        for pos in position:
            if isinstance(pos, str):
                result.append(pos)
            elif len(pos) == 2:
                if isinstance(pos[0], int) and isinstance(pos[1], int):
                    result.append(get_column_letter(pos[0]+1)+str(pos[1]+1))
                else:
                    result.append(None)

        return result

    @staticmethod
    def get_position_coordinates(position):
        """
        Convert position string to coordinates.

        Parameters
        ----------
        position: str or (int, int)
            position string  ('A6', 'C4', ...)

        Returns
        -------
        (int, int)
            coordinates

        """
        if isinstance(position, str):
            match = re.match(r"([a-z]+)([0-9]+)", position, re.I)
            if match:
                items = match.groups()
            else:
                return None
            num = 0
            for c in items[0]:
                if c in string.ascii_letters:
                    num = num * 26 + (ord(c.upper()) - ord('A')) + 1
            coordinates = (num - 1, int(items[1]) - 1)
        elif len(position) == 2:
            coordinates = position
        else:
            coordinates = None
        return coordinates

    def get_spot(self, position, double=None):
        """
        Return image data for a specific spot.

        Parameters
        ----------
        position: str or (int, int)
            position string  ('A6', 'C4', ...) or coordinates
        double: bool
            double spot switch
        Returns
        -------
        dict
            Dictionary with the three data sets for the requested spot. ('foreground', 'background', 'raw')

        """

        if not self.is_finalized:
            warnings.warn('Data collection needs to be finalized to generate spot data.', RuntimeWarning)
            return None
        coordinates = self.get_position_coordinates(position)
        if coordinates is None:
            return None
        if double:
            coordinates = self._add_2d_tuple(coordinates, self.array_data['double_spot'])
        x_raw, y_raw = (self.grid_position[coordinates[1], coordinates[0]])
        result = dict()
        result['foreground'] = self.foregrounds[int(round(y_raw)):int(round(y_raw))+config.scale,
                                                int(round(x_raw)):int(round(x_raw))+config.scale, :]
        result['background'] = self.backgrounds[int(round(y_raw)):int(round(y_raw))+config.scale,
                                                int(round(x_raw)):int(round(x_raw))+config.scale, :]
        result['raw'] = self.raw_images[int(round(y_raw)):int(round(y_raw))+config.scale,
                                        int(round(x_raw)):int(round(x_raw))+config.scale, :]
        return result

    def light_reaction(self, t, c_enzyme):
        """
        Calculate the accumulated light intensity.

        Notes
        -----
        Function based on findings presented in "An Investigation of the Mechanism of the Luminescent Peroxidation
        of Luminol by Stopped Flow Techniques" **Milton J. Cormier and Philip M. Prichard** - *J. Biol. Chem.* 1968 243: 4706.


        Parameters
        ----------
        t: float
            time in seconds

        c_enzyme: float
            enzyme concentration in mol/L

        Returns
        -------
        float
            accumulated light intensity

        """
        return self._kappa * (np.exp(-self.k_reac * c_enzyme * self.start_time)
                              - np.exp(-self.k_reac * c_enzyme * (t + self.start_time)))

    def kappa_error(self, kappa):
        """
        Calculate the error for a given kappa for a subset of spots in `_fit_selection`.

        Parameters
        ----------
        kappa: float
            combined parameter

        Returns
        -------
        float
            absolute error sum over all selected samples

        """

        self._kappa = kappa
        err_sum = 0
        for spot_raw in self._fit_selection:
            popt, pcov = curve_fit(self.light_reaction, self.exposure, spot_raw, p0=1E-10)
            err_sum += sum([np.abs(self.light_reaction(t, popt[0]) - spot_raw[n]) for n, t in enumerate(self.exposure)])
        return err_sum

    def minimize_kappa(self):
        """
        Finds the optimal kappa for a subset of `kappa_fit_count` numbers of the brightest spots.

        """
        selection = []
        for entry in self.array_data['spots']:
            spot = self.get_spot(entry['position'])
            selection.append((entry['position'], np.max(spot['raw'])))
        self._fit_selection = [np.average(self.get_spot(entry[0])['raw'], (1, 0)) - self.bg_parameters
                               for entry in sorted(selection, key=lambda x: x[1], reverse=True)[:self.kappa_fit_count]]
        result = minimize_scalar(self.kappa_error, bounds=(1E-5, 1))
        self._kappa = result.x

    @staticmethod
    def evaluate_spots_raw(spot):
        return np.average(spot['raw'], (0, 1))

    def evaluate_spots_raw_bg_corrected(self, spot):
        return np.average(spot['raw'], (0, 1))-self.bg_parameters

    @staticmethod
    def evaluate_spot_local_bg(spot):
        ratio = spot['raw'] / spot['background']
        return np.mean(ratio), np.std(ratio)

    def evaluate_spot_hist_foreground(self, spot):
        averages = np.average(spot['foreground'], (0, 1))
        data = stats.linregress(self.bg_parameters, averages)
        return data[0], data[1], data[2]**2

    def evaluate_spot_hist_raw(self, spot):
        averages = np.average(spot['raw'], (0, 1))
        data = stats.linregress(self.bg_parameters, averages)
        return data[0], data[1], data[2]**2

    def evaluate_spot_reac(self, spot):
        """

        Parameters
        ----------
        spot: dict
            spot data from `get_spot()`

        Returns
        -------
        (float, float)
            enzyme concentration in mol/L and standard deviation

        """

        popt, pcov = curve_fit(self.light_reaction, self.exposure, np.average(spot['raw']-self.bg_parameters,
                                                                              (0, 1)), p0=1E-10)
        return popt[0], float(np.sqrt(pcov[0]))

    def evaluate(self, position=None, norm='hist_raw', just_value=False, double_spot=False):
        """
        Evaluate a spot or the complete array.

        Notes
        -----
        You can select between different evaluation modes, by setting `norm`.

        - raw: list of averages for all time-steps based on the original image
        - raw_bg: as raw but reduced by the histogram based background value
        - local_bg: mean of the ratios between the original images, and the extracted backgrounds
        - hist_fg: the linear correlation between background (histogram) evolution and the average foreground value
        - hist_raw: as hist_fg but compared to the original image
        - reac: estimate of the catalytic enzyme concentration

        The reaction based model is just available when the exposure times can be accessed in the loaded images.
        While resulting value is in mol/L it is important to note that it is just an estimation based on the reaction
        constants published in "An Investigation of the Mechanism of the Luminescent Peroxidation of Luminol by
        Stopped Flow Techniques" **Milton J. Cormier and Philip M. Prichard** - *J. Biol. Chem.* 1968 243: 4706.


        Parameters
        ----------
        position: str or (int, int)
            position string  ('A6', 'C4', ...) or coordinates
        norm: str
            evaluation strategy selection
        just_value: bool
            return only the values
        double_spot: bool
            no position > if True double spots will be averaged and just the final value is returned
            specific position > return also the value of the partner

        Returns
        -------
        dict or array_like
            if `just_value` an array of values is returned else a dictionary including spot information

        """

        if not self.is_finalized:
            warnings.warn('Data collection needs to be finalized to evaluate.', RuntimeWarning)
            return None

        if norm == 'raw':
            evaluate_spots = self.evaluate_spots_raw
        elif norm == 'raw_bg':
            evaluate_spots = self.evaluate_spots_raw_bg_corrected
        elif norm == 'local_bg':
            evaluate_spots = self.evaluate_spot_local_bg
        elif norm == 'hist_fg':
            evaluate_spots = self.evaluate_spot_hist_foreground
        elif norm == 'hist_raw':
            evaluate_spots = self.evaluate_spot_hist_raw
        elif norm == 'reac':
            if not self.has_exposure:
                self.verbose_print('The reaction model needs exposure time date. Please select a different mode.')
                return None
            evaluate_spots = self.evaluate_spot_reac
        else:
            return None
        data = []
        if position is None:
            for entry in self.array_data['spots']:
                spot = self.get_spot(entry['position'])
                value = evaluate_spots(spot)
                if isinstance(value, tuple):
                    ds_idx = 0
                else:
                    ds_idx = -1
                if 'double_spot' in self.array_data:
                    ds_spot = self.get_spot(entry['position'], double=self.array_data['double_spot'])
                    ds_value = evaluate_spots(ds_spot)
                    if just_value:
                        if double_spot:
                            data.append(float((value[ds_idx] + ds_value[ds_idx])/2))
                        else:
                            data.append(value[ds_idx])
                            data.append(ds_value[ds_idx])
                    else:
                        if double_spot:
                            data.append(dict(position=[entry['position'], [entry['position'][0] +
                                                                           self.array_data['double_spot'][0],
                                                                           entry['position'][1] +
                                                                           self.array_data['double_spot'][1]]],
                                             info=entry['info'],
                                             value=float((value[ds_idx] + ds_value[ds_idx]) / 2)))
                        else:
                            data.append(dict(position=entry['position'], info=entry['info'], value=value))
                            position = [entry['position'][0] + self.array_data['double_spot'][0],
                                        entry['position'][1] + self.array_data['double_spot'][1]]
                            data.append(dict(position=position, info=entry['info'], value=ds_value))
                else:
                    if just_value:
                        data.append(value[ds_idx])
                    else:
                        data.append(dict(position=entry['position'], info=entry['info'], value=value))

        else:
            position = self.get_position_coordinates(position)
            spot = self.get_spot(position)
            value = evaluate_spots(spot)
            info = ''
            for entry in self.array_data['spots']:
                if list(entry['position']) == list(position):
                    info = entry['info']
                    break
                if (entry['position'][0] == position[0] - self.array_data['double_spot'][0] and
                        entry['position'][1] == position[1] - self.array_data['double_spot'][1]):
                    info = entry['info']
                    break

            if just_value:
                data.append(value)
            else:
                data.append(dict(position=position, info=info, value=value))
            if 'double_spot' in self.array_data and double_spot:
                spot = self.get_spot(position, double=True)
                value = evaluate_spots(spot)
                if just_value:
                    data.append(value)
                else:
                    position = (position[0] + self.array_data['double_spot'][0],
                                position[1] + self.array_data['double_spot'][1])
                    data.append(dict(position=position, info=info, value=value))

        return data

    def figure_reaction_fit(self, file=None, count=None):
        """
        Combine an array time series into an overview image.

        Parameters
        ----------
        file:
            file can be a path to a file (a string), a file-like object, or a path-like object (default svg);
            if file is None result is directly shown
        count: int
            number of the brightest spots to include in the plot
        """

        if not self.is_finalized:
            warnings.warn('Data collection needs to be finalized to generate the figure.', RuntimeWarning)
            return None

        if count is None:
            count = self.kappa_fit_count

        selection = []
        for entry in self.array_data['spots']:
            spot = self.get_spot(entry['position'])
            selection.append((entry['position'], np.max(spot['raw'])))
        fit_selection = [entry[0] for entry in sorted(selection, key=lambda x: x[1], reverse=True)[:count]]

        fig, ax = plt.subplots()
        ax.set_title(r'Determine $\kappa$')
        ax.set_ylabel('light flux $L$')
        ax.set_xlabel('exposure time $t$ (min)')
        for position in fit_selection:
            spot = self.get_spot(position)
            y_raw = self.evaluate_spots_raw_bg_corrected(spot)
            c_enzyme = self.evaluate_spot_reac(spot)[0]
            self.verbose_print(c_enzyme)
            y = []
            for t in self.exposure:
                y.append(self.light_reaction(t, c_enzyme))
            ax.plot(self.exposure/60, y, c='black')
            ax.scatter(self.exposure/60, y_raw, c='#469DFF', s=3, zorder=5)

        if file is None:  # pragma: no cover
            fig.show()
        else:
            if isinstance(file, os.PathLike) or isinstance(file, str):
                fig.savefig(file)
            elif isinstance(file, (io.RawIOBase, io.BufferedIOBase)):
                fig.savefig(file, format='svg')

    def figure_alignment(self, kind='raw', file=None, example=-1, max_size=None):
        """
        Build the array from the extracted spots in a chess pattern to check alignment.

        Parameters
        ----------
        kind: str
            raw: raw content; bg: background; fg: foreground
        file:
            file can be a path to a file (a string), a file-like object, or a path-like object;
            if file is None result is directly shown
        example: int
            what image to use
        max_size: int
            size of the longest edge in pixels
        """

        if not self.is_finalized:
            warnings.warn('Data collection needs to be finalized to generate the figure.', RuntimeWarning)
            return None

        x_num = sum(self.array_data['net_layout_x'])
        y_num = sum(self.array_data['net_layout_y'])
        x_length = sum(self.array_data['net_layout_x']) * config.scale
        y_length = sum(self.array_data['net_layout_y']) * config.scale
        align_image = np.zeros((y_length, x_length))
        data_match = {'raw': 'raw', 'fg': 'foreground', 'bg': 'background'}
        image_match = {'raw': 'raw_images', 'fg': 'foregrounds', 'bg': 'backgrounds'}
        full_image = np.log(getattr(self, image_match[kind]) + 0.1)
        min_full, max_full = np.min(full_image), np.max(full_image)
        for x in range(y_num):
            for y in range(x_num):
                spot = (self.get_spot((x, y))[data_match[kind]])[:, :, example]
                spot = np.log(spot + 0.1)
                spot = ((spot - min_full) * (1 / (max_full - min_full)))

                x_pos = [x * config.scale, (x + 1) * config.scale]
                y_pos = [y * config.scale, (y + 1) * config.scale]
                if (x+y) % 2:
                    align_image[x_pos[0]:x_pos[1], y_pos[0]:y_pos[1]] = spot
                else:
                    align_image[x_pos[0]:x_pos[1], y_pos[0]:y_pos[1]] = invert(spot)

        if file is None:  # pragma: no cover
            fig, ax = plt.subplots()
            ax.set_title(f'Align test {kind}')
            ax.imshow(align_image, interpolation='nearest', cmap=plt.cm.CMRmap, vmin=0, vmax=1)
            ax.axes.get_xaxis().set_visible(False)
            ax.axes.get_yaxis().set_visible(False)
            fig.show()

        else:
            image = plt.cm.CMRmap(align_image)[:, :, :3]
            if max_size:
                factor = max_size / np.max(image.shape)
                if factor < 1:
                    image = rescale(image, factor)

            image = img_as_ubyte(image)
            if isinstance(file, os.PathLike) or isinstance(file, str):
                ski_io.imsave(str(Path(file).absolute()), image)
            elif isinstance(file, (io.RawIOBase, io.BufferedIOBase)):
                im = Image.fromarray(image)
                im.save(file, format='JPEG')

    def figure_contact_sheet(self, kind='raw', file=None, max_size=None):
        """
        Combine an array time series into an overview image.

        Parameters
        ----------
        kind: str
            raw: raw content; bg: background; fg: foreground
        file:
            file can be a path to a file (a string), a file-like object, or a path-like object;
            if file is None result is directly shown
        max_size: int
            size of the longest edge in pixels
        """

        if not self.is_finalized:
            warnings.warn('Data collection needs to be finalized to generate the figure.', RuntimeWarning)
            return None
        pad = 20
        y_count = int(np.ceil(np.sqrt(self.raw_images.size) / self.raw_images.shape[1]))
        x_count = int(np.ceil(self.raw_images.shape[2]/y_count))
        y_max = int(y_count * self.raw_images.shape[1] + (y_count-1) * pad)
        x_max = int(x_count * self.raw_images.shape[0] + (x_count-1) * pad)

        data_match = {'raw': 'raw_images', 'fg': 'foregrounds', 'bg': 'backgrounds'}
        image = np.ones((x_max, y_max))

        for x_c in range(x_count):
            for y_c in range(y_count):
                x = x_c * pad + x_c * self.raw_images.shape[0]
                y = y_c * pad + y_c * self.raw_images.shape[1]
                try:
                    image[x:x + self.raw_images.shape[0], y:y + self.raw_images.shape[1]] = (
                        getattr(self, data_match[kind])[:, :, x_c * y_count + y_c])
                except IndexError:
                    pass

        image = np.log(image+0.1)
        image = ((image - np.min(image)) * (1 / (np.max(image) - np.min(image))))

        if file is None: # pragma: no cover
            fig, ax = plt.subplots()
            ax.set_title(f'Contact sheet {kind}')
            ax.imshow(image, interpolation='nearest', cmap=plt.cm.CMRmap_r, vmin=0, vmax=1)
            ax.axes.get_xaxis().set_visible(False)
            ax.axes.get_yaxis().set_visible(False)
            fig.show()

        else:
            image = plt.cm.CMRmap_r(image)[:, :, :3]
            if max_size:
                factor = max_size / np.max(image.shape)
                if factor < 1:
                    image = rescale(image, factor)

            image = img_as_ubyte(image)
            if isinstance(file, os.PathLike) or isinstance(file, str):
                ski_io.imsave(str(Path(file).absolute()), image)
            elif isinstance(file, (io.RawIOBase, io.BufferedIOBase)):
                im = Image.fromarray(image)
                im.save(file, format='JPEG')

    def figure_contact_sheet_spot(self, position, kind='raw', file=None, max_size=None):
        """
        Combine an array time series into an overview image.

        Parameters
        ----------
        position: str or (int, int)
            position string  ('A6', 'C4', ...) or coordinates
        kind: str
            raw: raw content; bg: background; fg: foreground
        file:
            file can be a path to a file (a string), a file-like object, or a path-like object;
            if file is None result is directly shown
        max_size: int
            size of the longest edge in pixels
        """

        if not self.is_finalized:
            warnings.warn('Data collection needs to be finalized to generate the figure.', RuntimeWarning)
            return None

        data_match = {'raw': 'raw', 'fg': 'foreground', 'bg': 'background'}

        spot = self.get_spot(position)
        pad = 2
        sx, sy, spot_count = spot['raw'].shape
        y_count = int(np.ceil(np.sqrt(spot_count)))
        x_count = int(np.ceil(spot_count / y_count))
        y_max = int(y_count * sy + (y_count - 1) * pad)
        x_max = int(x_count * sx + (x_count - 1) * pad)

        css = np.ones((x_max, y_max))
        for x_c in range(x_count):
            for y_c in range(y_count):
                x = x_c * pad + x_c * sx
                y = y_c * pad + y_c * sy
                try:
                    css[x:x + sx, y:y + sy] = spot[data_match[kind]][:, :, x_c * y_count + y_c]
                except IndexError:
                    pass

        image = np.log(css + 0.1)
        image = ((image - np.min(image)) * (1 / (np.max(image) - np.min(image))))

        if file is None: # pragma: no cover
            fig, ax = plt.subplots()
            ax.set_title(f'Contact spot sheet ({data_match[kind]})')
            ax.imshow(image, interpolation='nearest', cmap=plt.cm.CMRmap_r, vmin=0, vmax=1)
            ax.axes.get_xaxis().set_visible(False)
            ax.axes.get_yaxis().set_visible(False)
            fig.show()

        else:
            image = plt.cm.CMRmap_r(image)[:, :, :3]
            if max_size:
                factor = max_size / np.max(image.shape)
                if factor < 1:
                    image = rescale(image, factor)

            image = img_as_ubyte(image)
            if isinstance(file, os.PathLike) or isinstance(file, str):
                ski_io.imsave(str(Path(file).absolute()), image)
            elif isinstance(file, (io.RawIOBase, io.BufferedIOBase)):
                im = Image.fromarray(image)
                im.save(file, format='JPEG')

    def report(self, file=None, norm='hist_raw', report_type=None, additional_info=None):
        """
        Summarize the results for a specific evaluation strategy in a report.

        Notes
        -----
        Different report modules are available:

        - json
        - csv (no metadata are stored in this format!)
        - excel

        Parameters
        ----------
        file:
            can be a path to a file (a string), a path-like object, or a file-like object
        norm: str
            evaluation strategy selection (see ArrayAnalyse.evaluate)
        report_type: str
            set the report type, if none try to guess depending on file name
        """

        if not self.is_finalized:
            warnings.warn('Data collection needs to be finalized to create a report.', RuntimeWarning)
            return None

        report_types = {
            'json': dict(suffix=('.json',), func=Report.exp_json),
            'csv': dict(suffix=('.csv', '.txt', ''), func=Report.exp_csv),
            'excel': dict(suffix=('.xlsx',), func=Report.exp_excel),
            'latex': dict(suffix=('.tex',), func=Report.exp_latex)
        }

        if report_type is None:
            if isinstance(file, os.PathLike) or isinstance(file, str):
                file = Path(file)
                if file.suffix.lower() == '.xls':
                    file = file.with_suffix('.xlsx')
                for key, entry in report_types.items():
                    if file.suffix.lower() in entry['suffix']:
                        report_type = key
                        break

        if report_type not in report_types:
            warnings.warn(f'"{report_type}" is not defined as report type.', RuntimeWarning)
            return

        if file is None:
            warnings.warn(f'No file was given.', RuntimeWarning)
            return

        report_types[report_type]['func'](aa=self, file=file, norm=norm, additional_info=additional_info)

